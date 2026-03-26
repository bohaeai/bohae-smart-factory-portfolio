from __future__ import annotations

from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


class ExcelWriter:
    def __init__(self, out_path: str):
        self.out_path = out_path

    def write(self, result: Dict[str, Any], data: Optional[Dict[str, Any]] = None) -> str:
        sheets: Dict[str, List[Dict[str, Any]]] = {
            "META": result.get("meta_rows") or [],
            "PLAN_DEMAND": result.get("plan_rows") or [],
            "PLAN_SEGMENT": result.get("seg_rows") or [],
            "SPLIT_DETAIL": result.get("split_rows") or [],
            "CHANGEOVER_AUDIT": result.get("changeover_rows") or [],
            "PLAN_STAFF": result.get("staff_rows") or [],
            "POLICY_AUDIT": result.get("policy_rows") or [],
            "BREAK_SCHEDULE": result.get("break_rows") or [],
            "INFEASIBLE_DEMAND": result.get("infeasible_rows") or [],
            "FILTER_TRACE": result.get("filter_trace_rows") or [],
            "LINE_CANDIDATES": result.get("line_candidates_rows") or [],
            "QC_VALIDATION": result.get("qc_rows") or [],
            "OBJECTIVE_VALUES": result.get("objective_rows") or [],
            "SCORE_BREAKDOWN": result.get("score_rows") or [],
            "SLACK_ANALYSIS": result.get("slack_rows") or [],
            "UTILIZATION_HEATMAP": result.get("util_rows") or [],
            "DECISION_LOG": result.get("decision_log_rows") or [],
            "DATA_QUALITY": result.get("data_quality_rows") or [],
            "SSOT_ISSUE": result.get("ssot_issue_rows") or [],
            "SOLVER_STATS": result.get("solver_stats_rows") or [],
            "PASS_STATS": result.get("pass_stats_rows") or [],
            "STAFF_UTILIZATION": result.get("staff_util_rows") or [],
        }

        if data is not None:
            sheets["SHEET_REGISTRY"] = data.get("sheet_registry") or []
        else:
            # allow result to include registry
            sheets["SHEET_REGISTRY"] = result.get("sheet_registry") or []

        # Trace sheet (flattened)
        trace = result.get("trace") or {}
        trace_rows = []
        for k, v in trace.items():
            trace_rows.append({"KEY": str(k), "VALUE": str(v)})
        sheets["TRACE"] = trace_rows
        sheets["RUN_SUMMARY_KO"] = self._build_run_summary_rows(result)
        sheets["GUIDE_KO"] = self._build_guide_rows()

        with pd.ExcelWriter(self.out_path, engine="openpyxl") as xw:
            for name, rows in sheets.items():
                df = pd.DataFrame(rows)
                df.to_excel(xw, sheet_name=name, index=False)

        self._apply_readability_format()
        return self.out_path

    def _build_run_summary_rows(self, result: Dict[str, Any]) -> List[Dict[str, Any]]:
        meta_map: Dict[str, str] = {}
        for row in (result.get("meta_rows") or []):
            key = str(row.get("KEY", "")).strip()
            if key:
                meta_map[key] = str(row.get("VALUE", "")).strip()

        trace_map: Dict[str, str] = {}
        for key, value in (result.get("trace") or {}).items():
            trace_map[str(key).strip()] = str(value).strip()

        demand_rows = result.get("demand_rows") or result.get("plan_rows") or []
        demand_df = pd.DataFrame(demand_rows)
        demand_uns_cnt = ""
        demand_uns_qty = ""
        demand_top_uns = ""
        if not demand_df.empty and "IS_SCHEDULED" in demand_df.columns:
            scheduled = demand_df["IS_SCHEDULED"].map(lambda x: bool(x) if isinstance(x, bool) else str(x).strip().lower() in {"1", "true", "y", "yes"})
            uns_df = demand_df.loc[~scheduled].copy()
            demand_uns_cnt = str(int(len(uns_df)))
            if "ORDER_QTY" in uns_df.columns:
                demand_uns_qty = str(int(pd.to_numeric(uns_df["ORDER_QTY"], errors="coerce").fillna(0).sum()))
            if not uns_df.empty:
                label_col = "PRODUCT_NAME_KO" if "PRODUCT_NAME_KO" in uns_df.columns else ("PRODUCT_ID" if "PRODUCT_ID" in uns_df.columns else "")
                if label_col:
                    counts = uns_df[label_col].fillna("").astype(str).str.strip()
                    counts = counts[counts != ""].value_counts()
                    if not counts.empty:
                        demand_top_uns = f"{counts.index[0]} ({int(counts.iloc[0])}건)"

        objective_rows = result.get("objective_rows") or []
        unscheduled_count = self._extract_objective_value(objective_rows, ["UNSCHEDULED_COUNT", "objective_UNSCHEDULED_COUNT"])
        unscheduled_qty = self._extract_objective_value(objective_rows, ["UNSCHEDULED_QTY", "objective_UNSCHEDULED_QTY"])
        if not unscheduled_count:
            unscheduled_count = trace_map.get("objective_UNSCHEDULED_COUNT", "")
        if not unscheduled_qty:
            unscheduled_qty = trace_map.get("objective_UNSCHEDULED_QTY", "")
        if demand_uns_cnt != "":
            unscheduled_count = demand_uns_cnt
        if demand_uns_qty != "":
            unscheduled_qty = demand_uns_qty

        infeasible_rows = result.get("infeasible_rows") or []
        top_infeasible = ""
        if infeasible_rows:
            counts: Dict[str, int] = {}
            for row in infeasible_rows:
                label = str(row.get("PRODUCT_NAME_KO") or row.get("PRODUCT_ID") or "").strip()
                if not label:
                    continue
                counts[label] = counts.get(label, 0) + 1
            if counts:
                top_name = max(counts, key=counts.get)
                top_infeasible = f"{top_name} ({counts[top_name]}건)"
        if not top_infeasible and demand_top_uns:
            top_infeasible = demand_top_uns

        solver_status = meta_map.get("SOLVER_STATUS", "") or trace_map.get("solve_status", "") or trace_map.get("status", "")

        rows: List[Dict[str, Any]] = [
            {"구분": "실행 식별자", "값": meta_map.get("RUN_ID", ""), "설명": "이번 계산 실행 ID"},
            {"구분": "시나리오", "값": meta_map.get("SCENARIO", ""), "설명": "적용된 계획 시나리오"},
            {"구분": "계획 기간 시작", "값": meta_map.get("START_DATE", ""), "설명": "계획 시작일"},
            {"구분": "계획 기간 종료", "값": meta_map.get("END_DATE", ""), "설명": "계획 종료일"},
            {"구분": "솔버 상태", "값": solver_status, "설명": "OPTIMAL/FEASIBLE/FAIL"},
            {"구분": "선택 패스", "값": trace_map.get("selected_pass", ""), "설명": "2단계/폴백 패스 정보"},
            {"구분": "미배정 건수", "값": unscheduled_count, "설명": "0이면 전수요 배정"},
            {"구분": "미배정 수량", "값": unscheduled_qty, "설명": "병 단위 기준"},
            {"구분": "주요 미배정 품목", "값": top_infeasible, "설명": "가장 자주 미배정된 품목"},
            {"구분": "권장 확인 시트", "값": "PLAN_SEGMENT / INFEASIBLE_DEMAND / SOLVER_STATS", "설명": "원인 진단 핵심 시트"},
        ]
        return rows

    @staticmethod
    def _extract_objective_value(rows: List[Dict[str, Any]], keys: List[str]) -> str:
        for row in rows:
            row_key = str(row.get("OBJECTIVE") or row.get("KEY") or row.get("NAME") or "").strip()
            if row_key in keys:
                return str(row.get("VALUE", "")).strip()
        return ""

    @staticmethod
    def _build_guide_rows() -> List[Dict[str, Any]]:
        return [
            {"목적": "사람이 먼저 봐야 할 파일", "설명": "*_ops.xlsx(현장 운영용), *_rich.xlsx(감사/리포트용)부터 확인", "비고": "본 파일은 엔진 상세 데이터 포함"},
            {"목적": "실행 요약 확인", "설명": "RUN_SUMMARY_KO 시트에서 상태/미배정/핵심 원인 우선 확인", "비고": "미배정이 0이 아니면 INFEASIBLE_DEMAND 점검"},
            {"목적": "배정 결과", "설명": "PLAN_SEGMENT에서 라인/시간/수량 기준 실제 배정 확인", "비고": "필터 기준: LINE_ID, WORK_DATE"},
            {"목적": "미배정 원인", "설명": "INFEASIBLE_DEMAND + FILTER_TRACE + LINE_CANDIDATES 교차 확인", "비고": "SSOT 결함은 SSOT_ISSUE 참고"},
            {"목적": "성능/품질", "설명": "SOLVER_STATS, PASS_STATS, DATA_QUALITY 시트 확인", "비고": "fallback 사용 여부를 먼저 확인"},
        ]

    def _apply_readability_format(self) -> None:
        try:
            wb = load_workbook(self.out_path)
        except Exception:
            return

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(bold=True, color="FFFFFF")
        header_align = Alignment(vertical="center", wrap_text=True)
        body_align = Alignment(vertical="top", wrap_text=True)

        for ws in wb.worksheets:
            max_row = int(ws.max_row or 0)
            max_col = int(ws.max_column or 0)
            if max_row <= 0 or max_col <= 0:
                continue

            for cell in ws[1]:
                if cell.value is None:
                    continue
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align

            if max_row >= 2:
                ws.freeze_panes = "A2"
                ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
            ws.sheet_view.zoomScale = 110

            sample_last_row = min(max_row, 500)
            for row_idx in range(2, sample_last_row + 1):
                for col_idx in range(1, max_col + 1):
                    ws.cell(row=row_idx, column=col_idx).alignment = body_align

            for col_idx in range(1, max_col + 1):
                letter = get_column_letter(col_idx)
                max_len = 0
                for row_idx in range(1, sample_last_row + 1):
                    value = ws.cell(row=row_idx, column=col_idx).value
                    if value is None:
                        continue
                    length = len(str(value))
                    if length > max_len:
                        max_len = length
                    if max_len >= 64:
                        break
                ws.column_dimensions[letter].width = max(10, min(64, max_len + 2))

        try:
            wb.save(self.out_path)
        except Exception:
            return
