from __future__ import annotations

import json
import os
import re
from dataclasses import dataclass
from datetime import date
from typing import Any, Dict, Iterable, List, Literal, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..loaders.excel_io import load_sheet, normalize_cols
from ..utils.helpers import file_sha256, parse_date, parse_datetime, s, utcnow_iso
from .xlsx_pretty import ensure_sheet, write_df

QtyUnit = Literal["bottle", "case", "both"]

_HEADER_FILL = PatternFill("solid", fgColor="E7E6E6")
_WEEKEND_FILL = PatternFill("solid", fgColor="F3F3F3")
_HEADER_FONT = Font(bold=True)
_WRAP_TOP = Alignment(vertical="top", wrap_text=True)
_TOP = Alignment(vertical="top")
_CENTER = Alignment(horizontal="center", vertical="center")
_WEEKDAY_KO = ["월", "화", "수", "목", "금", "토", "일"]


@dataclass(frozen=True)
class _SsotMaps:
    product_name_by_id: Dict[str, str]
    line_name_by_id: Dict[str, str]
    line_active_by_id: Dict[str, bool]
    pack_qty_by_product_id: Dict[str, float]


def _truthy(x: Any) -> bool:
    return str(x).strip().upper() in {"Y", "1", "TRUE", "T"}


def _norm_id(x: Any) -> str:
    return s(x).upper()


def _load_json(path: Optional[str]) -> Dict[str, Any]:
    if not path:
        return {}
    p = str(path)
    if not os.path.exists(p):
        return {}
    try:
        return json.loads(open(p, "r", encoding="utf-8").read())
    except Exception:
        return {}


def _load_line_display_map(path: Optional[str]) -> Dict[str, str]:
    if not path:
        return {}
    p = str(path)
    if not os.path.exists(p):
        return {}
    txt = open(p, "r", encoding="utf-8").read()
    ext = os.path.splitext(p)[1].lower()

    if ext in {".json", ".jsn"}:
        try:
            payload = json.loads(txt)
        except Exception:
            return {}
        if isinstance(payload, dict):
            return {_norm_id(k): s(v) for k, v in payload.items() if s(k)}
        if isinstance(payload, list):
            out: Dict[str, str] = {}
            for row in payload:
                if not isinstance(row, dict):
                    continue
                key = _norm_id(row.get("line_id") or row.get("LINE_ID"))
                val = s(row.get("display_name") or row.get("DISPLAY_NAME") or row.get("label") or row.get("LABEL"))
                if key and val:
                    out[key] = val
            return out
        return {}

    # Lightweight YAML parser: `KEY: VALUE` lines only.
    out: Dict[str, str] = {}
    for raw in txt.splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        if ":" not in line:
            continue
        key, val = line.split(":", 1)
        key = _norm_id(key)
        val = s(val)
        if key and val:
            out[key] = val
    return out


def _sheet_by_prefix(xls: pd.ExcelFile, prefix: str) -> pd.DataFrame:
    df, _ = load_sheet(xls, prefix)
    return normalize_cols(df)


def _load_ssot_maps(ssot_xlsx_path: Optional[str]) -> _SsotMaps:
    if not ssot_xlsx_path or not os.path.exists(ssot_xlsx_path):
        return _SsotMaps({}, {}, {}, {})

    xls = pd.ExcelFile(ssot_xlsx_path)

    df10 = _sheet_by_prefix(xls, "10")
    df24 = _sheet_by_prefix(xls, "24")
    df25 = _sheet_by_prefix(xls, "25")
    df32 = _sheet_by_prefix(xls, "32")

    product_name_by_id: Dict[str, str] = {}
    if not df10.empty and "PRODUCT_ID" in df10.columns:
        name_col = "PRODUCT_NAME_KO" if "PRODUCT_NAME_KO" in df10.columns else ("ERP_PRODUCT_NAME_KO" if "ERP_PRODUCT_NAME_KO" in df10.columns else "")
        if name_col:
            t = df10[["PRODUCT_ID", name_col]].copy()
            t["PRODUCT_ID"] = t["PRODUCT_ID"].map(_norm_id)
            t[name_col] = t[name_col].map(s)
            t = t[t["PRODUCT_ID"].ne("")]
            product_name_by_id = dict(zip(t["PRODUCT_ID"], t[name_col]))

    line_name_by_id: Dict[str, str] = {}
    line_active_by_id: Dict[str, bool] = {}
    if not df32.empty and "LINE_ID" in df32.columns:
        name_col = "LINE_NAME_KO" if "LINE_NAME_KO" in df32.columns else ""
        t = df32.copy()
        t["LINE_ID"] = t["LINE_ID"].map(_norm_id)
        t = t[t["LINE_ID"].ne("")]
        if name_col:
            line_name_by_id = dict(zip(t["LINE_ID"], t[name_col].map(s)))
        if "IS_ACTIVE" in t.columns:
            line_active_by_id = dict(zip(t["LINE_ID"], t["IS_ACTIVE"].map(_truthy)))

    pack_qty_by_product_id: Dict[str, float] = {}
    if not df10.empty and not df25.empty and not df24.empty and "PACK_STYLE_ID" in df10.columns:
        prod = df10[["PRODUCT_ID", "PACK_STYLE_ID"]].copy()
        ps = df25[[c for c in ["PACK_STYLE_ID", "CASE_ID"] if c in df25.columns]].copy()
        case = df24[[c for c in ["CASE_ID", "PACK_QTY"] if c in df24.columns]].copy()
        if {"PACK_STYLE_ID", "CASE_ID"}.issubset(set(ps.columns)) and {"CASE_ID", "PACK_QTY"}.issubset(set(case.columns)):
            prod["PRODUCT_ID"] = prod["PRODUCT_ID"].map(_norm_id)
            prod["PACK_STYLE_ID"] = prod["PACK_STYLE_ID"].map(_norm_id)
            ps["PACK_STYLE_ID"] = ps["PACK_STYLE_ID"].map(_norm_id)
            ps["CASE_ID"] = ps["CASE_ID"].map(_norm_id)
            case["CASE_ID"] = case["CASE_ID"].map(_norm_id)
            case["PACK_QTY"] = pd.to_numeric(case["PACK_QTY"], errors="coerce")
            m = prod.merge(ps, on="PACK_STYLE_ID", how="left").merge(case, on="CASE_ID", how="left")
            m = m[pd.to_numeric(m["PACK_QTY"], errors="coerce").fillna(0) > 0]
            if not m.empty:
                m = m.drop_duplicates(subset=["PRODUCT_ID"], keep="first")
                pack_qty_by_product_id = dict(zip(m["PRODUCT_ID"], m["PACK_QTY"].astype(float)))

    return _SsotMaps(
        product_name_by_id=product_name_by_id,
        line_name_by_id=line_name_by_id,
        line_active_by_id=line_active_by_id,
        pack_qty_by_product_id=pack_qty_by_product_id,
    )


def _default_line_display(line_id: str, line_name_ko: str) -> str:
    lid = _norm_id(line_id)
    ln = s(line_name_ko)
    ln_u = ln.upper()
    if lid.endswith("_01") and "제조 1동 1호" in ln:
        return "1호"
    if lid.endswith("_02") and "제조 1동 2호" in ln:
        return "2호(12+2)"
    if lid.endswith("_03") and "제조 1동 3호" in ln:
        return "3호(12+2)"
    if lid.endswith("_05") and "제조 1동 5호" in ln:
        return "5호(15)"
    if "PET_A" in lid or "PET A" in ln_u:
        return "PET-A(5)"
    if "PET_B" in lid or "PET B" in ln_u:
        return "PET-B(7)"
    if ("B3" in lid and "캔" in ln) or ("CAN" in ln_u):
        return "CAN라인(1000)"
    if "B2" in lid or "제조 2동" in ln:
        return "탁주(8) 제조2동"
    return ln or line_id


def _line_sort_rank(display_name: str) -> Tuple[int, str]:
    nm = s(display_name).upper()
    if nm.startswith("1호".upper()):
        return (1, nm)
    if nm.startswith("2호".upper()):
        return (2, nm)
    if nm.startswith("3호".upper()):
        return (3, nm)
    if nm.startswith("5호".upper()):
        return (4, nm)
    if nm.startswith("PET-A"):
        return (5, nm)
    if nm.startswith("PET-B"):
        return (6, nm)
    if "탁주" in nm:
        return (7, nm)
    if "CAN" in nm:
        return (8, nm)
    return (99, nm)


def _fmt_date_ko(d: date) -> str:
    return f"{d.month}/{d.day}({_WEEKDAY_KO[d.weekday()]})"


def _fmt_int(x: float) -> str:
    return f"{int(round(float(x))):,}"


def _fmt_qty(qty: float, *, unit: QtyUnit, pack_qty: Optional[float]) -> str:
    q = float(qty)
    if unit == "bottle":
        return f"{_fmt_int(q)}본"
    if unit == "case":
        if pack_qty and pack_qty > 0:
            case_q = q / float(pack_qty)
            if abs(case_q - round(case_q)) < 1e-9:
                return f"{_fmt_int(round(case_q))}c/s"
            return f"{case_q:,.1f}c/s"
        return f"{_fmt_int(q)}본"
    # both
    if pack_qty and pack_qty > 0:
        case_q = q / float(pack_qty)
        if abs(case_q - round(case_q)) < 1e-9:
            c = _fmt_int(round(case_q))
        else:
            c = f"{case_q:,.1f}"
        return f"{_fmt_int(q)}본 ({c}c/s)"
    return f"{_fmt_int(q)}본"


def _meta_map_from_plan(xls: pd.ExcelFile) -> Dict[str, str]:
    if "META" not in xls.sheet_names:
        return {}
    df = normalize_cols(pd.read_excel(xls, sheet_name="META"))
    if not {"KEY", "VALUE"}.issubset(set(df.columns)):
        return {}
    t = df[["KEY", "VALUE"]].copy()
    t["KEY"] = t["KEY"].map(lambda v: s(v).upper())
    t["VALUE"] = t["VALUE"].map(s)
    t = t[t["KEY"].ne("")]
    return dict(zip(t["KEY"], t["VALUE"]))


def _read_required_sheets(plan_xlsx_path: str) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, Dict[str, str]]:
    xls = pd.ExcelFile(plan_xlsx_path)
    required = ["PLAN_SEGMENT", "SPLIT_DETAIL", "PLAN_DEMAND"]
    missing = [nm for nm in required if nm not in xls.sheet_names]
    if missing:
        raise ValueError(f"OPS_REPORT_MISSING_SHEETS: {','.join(missing)}")

    seg = normalize_cols(pd.read_excel(xls, sheet_name="PLAN_SEGMENT"))
    split = normalize_cols(pd.read_excel(xls, sheet_name="SPLIT_DETAIL"))
    dem = normalize_cols(pd.read_excel(xls, sheet_name="PLAN_DEMAND"))
    staff = normalize_cols(pd.read_excel(xls, sheet_name="PLAN_STAFF")) if "PLAN_STAFF" in xls.sheet_names else pd.DataFrame()
    meta = _meta_map_from_plan(xls)
    return seg, split, dem, staff, meta


def _build_seg_detail(seg: pd.DataFrame, split: pd.DataFrame, dem: pd.DataFrame) -> pd.DataFrame:
    seg_required = ["SEGMENT_ID", "DEMAND_ID", "PRODUCT_ID", "LINE_ID", "WORK_DATE", "START_DT", "END_DT", "DUR_MIN"]
    for col in seg_required:
        if col not in seg.columns:
            if seg.empty:
                seg[col] = pd.Series(dtype="object")
            else:
                raise ValueError(f"OPS_REPORT_PLAN_SEGMENT_MISSING_COL:{col}")
    if "SEGMENT_ID" not in split.columns:
        if split.empty:
            split["SEGMENT_ID"] = pd.Series(dtype="object")
        else:
            raise ValueError("OPS_REPORT_SPLIT_DETAIL_MISSING_COL:SEGMENT_ID/SEG_QTY")
    if "SEG_QTY" not in split.columns:
        if split.empty:
            split["SEG_QTY"] = pd.Series(dtype="float")
        else:
            raise ValueError("OPS_REPORT_SPLIT_DETAIL_MISSING_COL:SEGMENT_ID/SEG_QTY")

    sd = split.copy()
    sd["SEGMENT_ID"] = sd["SEGMENT_ID"].map(_norm_id)
    if "DEMAND_ID" in sd.columns:
        sd["DEMAND_ID"] = sd["DEMAND_ID"].map(_norm_id)
        grp_cols = ["SEGMENT_ID", "DEMAND_ID"]
    else:
        grp_cols = ["SEGMENT_ID"]
    sd["SEG_QTY"] = pd.to_numeric(sd["SEG_QTY"], errors="coerce")
    sd = sd.groupby(grp_cols, dropna=False, as_index=False)["SEG_QTY"].sum()

    sg = seg.copy()
    sg["SEGMENT_ID"] = sg["SEGMENT_ID"].map(_norm_id)
    sg["DEMAND_ID"] = sg["DEMAND_ID"].map(_norm_id)
    sg["PRODUCT_ID"] = sg["PRODUCT_ID"].map(_norm_id)
    sg["LINE_ID"] = sg["LINE_ID"].map(_norm_id)
    sg["WORK_DATE"] = sg["WORK_DATE"].map(parse_date)
    sg["START_DT"] = sg["START_DT"].map(parse_datetime)
    sg["END_DT"] = sg["END_DT"].map(parse_datetime)
    sg["DUR_MIN"] = pd.to_numeric(sg["DUR_MIN"], errors="coerce").fillna(0).astype(int)
    missing_work_date = sg["WORK_DATE"].isna() & sg["START_DT"].notna()
    if missing_work_date.any():
        sg.loc[missing_work_date, "WORK_DATE"] = sg.loc[missing_work_date, "START_DT"].map(lambda x: x.date() if x else None)

    join_cols = ["SEGMENT_ID", "DEMAND_ID"] if "DEMAND_ID" in sd.columns else ["SEGMENT_ID"]
    out = sg.merge(sd, on=join_cols, how="left")
    if out["SEG_QTY"].isna().any():
        miss = int(out["SEG_QTY"].isna().sum())
        raise ValueError(f"OPS_REPORT_SEG_QTY_MISSING:{miss}")

    dcols = [c for c in ["DEMAND_ID", "ORDER_QTY", "DUE_DATE", "IS_SCHEDULED", "TARDINESS_MIN"] if c in dem.columns]
    if dcols:
        dm = dem[dcols].copy()
        dm["DEMAND_ID"] = dm["DEMAND_ID"].map(_norm_id)
        if "ORDER_QTY" in dm.columns:
            dm["ORDER_QTY"] = pd.to_numeric(dm["ORDER_QTY"], errors="coerce").fillna(0.0)
        if "DUE_DATE" in dm.columns:
            dm["DUE_DATE"] = dm["DUE_DATE"].map(parse_date)
        if "TARDINESS_MIN" in dm.columns:
            dm["TARDINESS_MIN"] = pd.to_numeric(dm["TARDINESS_MIN"], errors="coerce").fillna(0.0)
        out = out.merge(dm, on="DEMAND_ID", how="left")

    return out


def _build_staff_headcount_by_date(staff: pd.DataFrame, seg_detail: pd.DataFrame) -> Dict[date, int]:
    if staff is None or staff.empty:
        return {}
    if "SEGMENT_ID" not in staff.columns:
        return {}

    sf = staff.copy()
    sf["SEGMENT_ID"] = sf["SEGMENT_ID"].map(_norm_id)
    seg_dates = seg_detail[["SEGMENT_ID", "WORK_DATE"]].drop_duplicates()
    sf = sf.merge(seg_dates, on="SEGMENT_ID", how="left")
    sf = sf[sf["WORK_DATE"].notna()]
    if sf.empty:
        return {}

    if "ASSIGN_STATUS" in sf.columns:
        sf = sf[sf["ASSIGN_STATUS"].map(lambda v: s(v).upper()).isin({"ASSIGNED", "", "OK"})]

    if "STAFF_ID" in sf.columns and sf["STAFF_ID"].map(s).ne("").any():
        sf["STAFF_ID"] = sf["STAFF_ID"].map(s)
        t = sf[sf["STAFF_ID"].ne("")].groupby("WORK_DATE", dropna=False)["STAFF_ID"].nunique()
        return {k: int(v) for k, v in t.to_dict().items() if pd.notna(k)}

    if "SLOT_IDX" in sf.columns:
        t = sf.groupby("WORK_DATE", dropna=False)["SLOT_IDX"].nunique()
        return {k: int(v) for k, v in t.to_dict().items() if pd.notna(k)}
    return {}


def _collect_line_columns(
    seg_detail: pd.DataFrame,
    ssot_maps: _SsotMaps,
    line_display_map: Dict[str, str],
    include_inactive_lines_in_columns: bool,
) -> Tuple[List[str], Dict[str, str]]:
    line_ids: List[str] = []
    seen: set[str] = set()

    if line_display_map:
        for lid in line_display_map.keys():
            if lid and lid not in seen:
                line_ids.append(lid)
                seen.add(lid)

    for lid in seg_detail["LINE_ID"].map(_norm_id).tolist():
        if lid and lid not in seen:
            line_ids.append(lid)
            seen.add(lid)

    # Add line master columns if requested.
    if include_inactive_lines_in_columns and ssot_maps.line_name_by_id:
        for lid in ssot_maps.line_name_by_id.keys():
            if lid not in seen:
                line_ids.append(lid)
                seen.add(lid)
    elif ssot_maps.line_name_by_id:
        for lid, active in ssot_maps.line_active_by_id.items():
            if not active:
                continue
            if lid not in seen:
                line_ids.append(lid)
                seen.add(lid)

    display_name_by_line: Dict[str, str] = {}
    for lid in line_ids:
        display_name_by_line[lid] = (
            line_display_map.get(lid)
            or _default_line_display(lid, ssot_maps.line_name_by_id.get(lid, ""))
        )

    map_order = {lid: idx for idx, lid in enumerate(line_display_map.keys())}
    line_ids = sorted(
        line_ids,
        key=lambda lid: (
            0 if lid in map_order else 1,
            map_order.get(lid, 9999),
            *_line_sort_rank(display_name_by_line.get(lid, lid)),
            lid,
        ),
    )
    return line_ids, display_name_by_line


def _date_window(seg_detail: pd.DataFrame, meta: Dict[str, str]) -> List[date]:
    start = parse_date(meta.get("START_DATE") or meta.get("HORIZON_START") or "")
    end = parse_date(meta.get("END_DATE") or meta.get("HORIZON_END") or "")
    if start and end and end >= start:
        return list(pd.date_range(start=start, end=end, freq="D").date)

    dmin = seg_detail["WORK_DATE"].dropna().min()
    dmax = seg_detail["WORK_DATE"].dropna().max()
    if dmin is None or dmax is None:
        return []
    return list(pd.date_range(start=dmin, end=dmax, freq="D").date)


def _build_grid_rows(
    seg_detail: pd.DataFrame,
    *,
    dates: Sequence[date],
    line_ids: Sequence[str],
    display_name_by_line: Dict[str, str],
    ssot_maps: _SsotMaps,
    max_cell_items: int,
    cell_qty_unit: QtyUnit,
    staff_by_date: Dict[date, int],
) -> Tuple[pd.DataFrame, float]:
    seg = seg_detail.copy()
    seg = seg.sort_values(["WORK_DATE", "LINE_ID", "START_DT", "END_DT"], kind="mergesort")

    cell_text_map: Dict[Tuple[date, str], str] = {}
    grid_qty_total = 0.0
    for (work_date, line_id), grp in seg.groupby(["WORK_DATE", "LINE_ID"], dropna=False):
        if pd.isna(work_date):
            continue
        g = grp.copy()
        g["SEG_QTY"] = pd.to_numeric(g["SEG_QTY"], errors="coerce").fillna(0.0)
        if "PRODUCT_NAME_KO" in g.columns:
            g["PRODUCT_NAME_SHOW"] = g["PRODUCT_NAME_KO"].map(s)
        else:
            g["PRODUCT_NAME_SHOW"] = ""
        missing_name = g["PRODUCT_NAME_SHOW"].eq("")
        if missing_name.any():
            g.loc[missing_name, "PRODUCT_NAME_SHOW"] = g.loc[missing_name, "PRODUCT_ID"].map(lambda x: ssot_maps.product_name_by_id.get(_norm_id(x), s(x)))
        missing_name = g["PRODUCT_NAME_SHOW"].eq("")
        if missing_name.any():
            g.loc[missing_name, "PRODUCT_NAME_SHOW"] = g.loc[missing_name, "PRODUCT_ID"].map(s)

        agg = (
            g.groupby(["PRODUCT_ID", "PRODUCT_NAME_SHOW"], dropna=False, as_index=False)
            .agg(TOTAL_QTY=("SEG_QTY", "sum"), FIRST_START=("START_DT", "min"))
            .sort_values(["FIRST_START", "PRODUCT_ID"], kind="mergesort")
        )
        grid_qty_total += float(agg["TOTAL_QTY"].sum())
        lines: List[str] = []
        for _, r in agg.iterrows():
            pid = _norm_id(r.get("PRODUCT_ID"))
            pnm = s(r.get("PRODUCT_NAME_SHOW")) or pid
            qty_txt = _fmt_qty(float(r.get("TOTAL_QTY", 0.0)), unit=cell_qty_unit, pack_qty=ssot_maps.pack_qty_by_product_id.get(pid))
            lines.append(f"{pnm}\n{qty_txt}")
        hidden = max(0, len(lines) - int(max_cell_items))
        if hidden > 0:
            lines = lines[: int(max_cell_items)] + [f"...(+{hidden} items)"]
        cell_text_map[(work_date, _norm_id(line_id))] = "\n".join(lines)

    rows: List[Dict[str, Any]] = []
    for d in dates:
        row: Dict[str, Any] = {"WORK_DATE": _fmt_date_ko(d)}
        for lid in line_ids:
            row[display_name_by_line.get(lid, lid)] = cell_text_map.get((d, lid), "")
        row["비고"] = ""
        staff_cnt = int(staff_by_date.get(d, 0))
        row["인원"] = staff_cnt if staff_cnt > 0 else ""
        row["암웨이출고"] = ""
        rows.append(row)

    return pd.DataFrame(rows), grid_qty_total


def _build_line_instruction(seg_detail: pd.DataFrame, ssot_maps: _SsotMaps, display_name_by_line: Dict[str, str], cell_qty_unit: QtyUnit) -> pd.DataFrame:
    df = seg_detail.copy().sort_values(["LINE_ID", "START_DT", "END_DT"], kind="mergesort")
    if df.empty:
        return pd.DataFrame(
            columns=[
                "WORK_DATE",
                "LINE_ID",
                "LINE_NAME_KO",
                "LINE_DISPLAY_NAME",
                "SEQ_IN_LINE",
                "START_DT",
                "END_DT",
                "DUR_MIN",
                "DEMAND_ID",
                "SEGMENT_ID",
                "PRODUCT_ID",
                "PRODUCT_NAME_KO",
                "SEG_QTY_BOTTLE",
                "SEG_QTY_CASE",
                "SETUP_IN_MIN",
                "DUE_DATE",
                "TARDINESS_MIN",
            ]
        )

    out = df.copy()
    out["SEQ_IN_LINE"] = out.groupby("LINE_ID", dropna=False).cumcount() + 1
    out["LINE_NAME_KO"] = out["LINE_ID"].map(lambda x: ssot_maps.line_name_by_id.get(_norm_id(x), ""))
    out["LINE_DISPLAY_NAME"] = out["LINE_ID"].map(lambda x: display_name_by_line.get(_norm_id(x), s(x)))
    if "PRODUCT_NAME_KO" in out.columns:
        out["PRODUCT_NAME_SHOW"] = out["PRODUCT_NAME_KO"].map(s)
    else:
        out["PRODUCT_NAME_SHOW"] = ""
    miss = out["PRODUCT_NAME_SHOW"].eq("")
    if miss.any():
        out.loc[miss, "PRODUCT_NAME_SHOW"] = out.loc[miss, "PRODUCT_ID"].map(lambda x: ssot_maps.product_name_by_id.get(_norm_id(x), s(x)))

    out["SEG_QTY_BOTTLE"] = pd.to_numeric(out["SEG_QTY"], errors="coerce").fillna(0.0)
    out["SEG_QTY_CASE"] = out.apply(
        lambda r: (
            (float(r["SEG_QTY_BOTTLE"]) / float(ssot_maps.pack_qty_by_product_id[_norm_id(r["PRODUCT_ID"])]))
            if _norm_id(r["PRODUCT_ID"]) in ssot_maps.pack_qty_by_product_id and ssot_maps.pack_qty_by_product_id[_norm_id(r["PRODUCT_ID"])] > 0
            else None
        ),
        axis=1,
    )

    setup_series = out["SETUP_IN_MIN"] if "SETUP_IN_MIN" in out.columns else pd.Series(index=out.index, dtype="float")
    tardiness_series = out["TARDINESS_MIN"] if "TARDINESS_MIN" in out.columns else pd.Series(index=out.index, dtype="float")
    keep = pd.DataFrame(
        {
            "WORK_DATE": out["WORK_DATE"],
            "LINE_ID": out["LINE_ID"],
            "LINE_NAME_KO": out["LINE_NAME_KO"],
            "LINE_DISPLAY_NAME": out["LINE_DISPLAY_NAME"],
            "SEQ_IN_LINE": out["SEQ_IN_LINE"],
            "START_DT": out["START_DT"],
            "END_DT": out["END_DT"],
            "DUR_MIN": out["DUR_MIN"],
            "DEMAND_ID": out["DEMAND_ID"],
            "SEGMENT_ID": out["SEGMENT_ID"],
            "PRODUCT_ID": out["PRODUCT_ID"],
            "PRODUCT_NAME_KO": out["PRODUCT_NAME_SHOW"],
            "SEG_QTY_BOTTLE": out["SEG_QTY_BOTTLE"],
            "SEG_QTY_CASE": out["SEG_QTY_CASE"],
            "SETUP_IN_MIN": pd.to_numeric(setup_series, errors="coerce").fillna(0).astype(int),
            "DUE_DATE": out.get("DUE_DATE"),
            "TARDINESS_MIN": pd.to_numeric(tardiness_series, errors="coerce").fillna(0.0),
        }
    )
    if cell_qty_unit == "case":
        keep["SEG_QTY_DISPLAY"] = keep["SEG_QTY_CASE"]
    elif cell_qty_unit == "both":
        keep["SEG_QTY_DISPLAY"] = keep.apply(
            lambda r: _fmt_qty(
                float(r["SEG_QTY_BOTTLE"]),
                unit="both",
                pack_qty=(None if pd.isna(r["SEG_QTY_CASE"]) else (float(r["SEG_QTY_BOTTLE"]) / float(r["SEG_QTY_CASE"])) if float(r["SEG_QTY_CASE"]) != 0 else None),
            ),
            axis=1,
        )
    else:
        keep["SEG_QTY_DISPLAY"] = keep["SEG_QTY_BOTTLE"]
    return keep


def _build_product_summary(seg_detail: pd.DataFrame, ssot_maps: _SsotMaps, plan_demand: pd.DataFrame) -> pd.DataFrame:
    seg = seg_detail.copy()
    seg["SEG_QTY"] = pd.to_numeric(seg["SEG_QTY"], errors="coerce").fillna(0.0)
    tardiness_series = seg["TARDINESS_MIN"] if "TARDINESS_MIN" in seg.columns else pd.Series(index=seg.index, dtype="float")
    seg["TARDINESS_MIN"] = pd.to_numeric(tardiness_series, errors="coerce").fillna(0.0)
    if seg.empty:
        return pd.DataFrame(
            columns=[
                "PRODUCT_ID",
                "PRODUCT_NAME_KO",
                "TOTAL_QTY_BOTTLE",
                "DEMAND_COUNT",
                "SEGMENT_COUNT",
                "LINE_USED_COUNT",
                "LINES_USED",
                "EARLIEST_START_DT",
                "LATEST_END_DT",
                "MIN_DUE_DATE",
                "MAX_DUE_DATE",
                "TOTAL_TARDINESS_MIN",
                "UNSCHEDULED_DEMAND_COUNT",
            ]
        )

    grp = seg.groupby("PRODUCT_ID", dropna=False)
    out = grp.agg(
        TOTAL_QTY_BOTTLE=("SEG_QTY", "sum"),
        DEMAND_COUNT=("DEMAND_ID", pd.Series.nunique),
        SEGMENT_COUNT=("SEGMENT_ID", "count"),
        LINE_USED_COUNT=("LINE_ID", pd.Series.nunique),
        EARLIEST_START_DT=("START_DT", "min"),
        LATEST_END_DT=("END_DT", "max"),
        MIN_DUE_DATE=("DUE_DATE", "min"),
        MAX_DUE_DATE=("DUE_DATE", "max"),
        TOTAL_TARDINESS_MIN=("TARDINESS_MIN", "sum"),
    ).reset_index()
    lines_top = (
        seg.groupby(["PRODUCT_ID", "LINE_ID"], dropna=False)["SEG_QTY"].sum().reset_index().sort_values(["PRODUCT_ID", "SEG_QTY"], ascending=[True, False], kind="mergesort")
    )
    line_name = {lid: _default_line_display(lid, ssot_maps.line_name_by_id.get(lid, "")) for lid in seg["LINE_ID"].map(_norm_id).unique()}
    top_map: Dict[str, str] = {}
    for pid, g in lines_top.groupby("PRODUCT_ID", dropna=False):
        top3 = [line_name.get(_norm_id(v), s(v)) for v in g["LINE_ID"].head(3).tolist()]
        top_map[_norm_id(pid)] = ", ".join([x for x in top3 if x])
    out["PRODUCT_ID"] = out["PRODUCT_ID"].map(_norm_id)
    out["PRODUCT_NAME_KO"] = out["PRODUCT_ID"].map(lambda x: ssot_maps.product_name_by_id.get(_norm_id(x), ""))
    out["LINES_USED"] = out["PRODUCT_ID"].map(lambda x: top_map.get(_norm_id(x), ""))

    uns_map: Dict[str, int] = {}
    if "IS_SCHEDULED" in plan_demand.columns and {"PRODUCT_ID", "DEMAND_ID"}.issubset(set(plan_demand.columns)):
        dm = plan_demand.copy()
        dm["PRODUCT_ID"] = dm["PRODUCT_ID"].map(_norm_id)
        dm["DEMAND_ID"] = dm["DEMAND_ID"].map(_norm_id)
        uns = dm[~dm["IS_SCHEDULED"].map(_truthy)].groupby("PRODUCT_ID", dropna=False)["DEMAND_ID"].nunique()
        uns_map = {k: int(v) for k, v in uns.to_dict().items()}
    out["UNSCHEDULED_DEMAND_COUNT"] = out["PRODUCT_ID"].map(lambda x: uns_map.get(_norm_id(x), 0)).astype(int)

    out = out.sort_values(["TOTAL_QTY_BOTTLE", "PRODUCT_ID"], ascending=[False, True], kind="mergesort")
    return out


def _write_grid_sheet(
    wb: Workbook,
    *,
    grid_df: pd.DataFrame,
    line_headers: Sequence[str],
    date_values: Sequence[date],
    index: int = 0,
) -> None:
    ws = ensure_sheet(wb, "01_주간생산계획", index=index)
    headers = list(grid_df.columns)

    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=str(h))
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER

    for r, row in enumerate(grid_df.itertuples(index=False), start=2):
        for c_idx, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=c_idx, value=val)
            c.alignment = _WRAP_TOP

        # weekend shading by underlying date list index
        idx = r - 2
        if 0 <= idx < len(date_values) and date_values[idx].weekday() >= 5:
            for c_idx in range(1, len(headers) + 1):
                ws.cell(row=r, column=c_idx).fill = _WEEKEND_FILL

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, len(grid_df) + 1)}"
    ws.sheet_view.zoomScale = 110

    # Widths
    ws.column_dimensions["A"].width = 13
    for idx in range(2, 2 + len(line_headers)):
        ws.column_dimensions[get_column_letter(idx)].width = 26
    tail_start = 2 + len(line_headers)
    if tail_start <= len(headers):
        ws.column_dimensions[get_column_letter(tail_start)].width = 18  # 비고
    if tail_start + 1 <= len(headers):
        ws.column_dimensions[get_column_letter(tail_start + 1)].width = 9  # 인원
    if tail_start + 2 <= len(headers):
        ws.column_dimensions[get_column_letter(tail_start + 2)].width = 13  # 암웨이출고

    # Row heights adapt to multiline
    for r in range(2, len(grid_df) + 2):
        max_lines = 1
        for c_idx in range(2, 2 + len(line_headers)):
            txt = s(ws.cell(row=r, column=c_idx).value)
            if txt:
                max_lines = max(max_lines, txt.count("\n") + 1)
        ws.row_dimensions[r].height = max(24, min(120, 16 + max_lines * 12))


def _apply_table_format(ws: Worksheet, datetime_cols: Iterable[str] = ()) -> None:
    ws.sheet_view.zoomScale = 110
    # header style
    for c_idx in range(1, ws.max_column + 1):
        c = ws.cell(row=1, column=c_idx)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER
    for r in range(2, ws.max_row + 1):
        for c_idx in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c_idx).alignment = _TOP

    hdr = {s(ws.cell(row=1, column=c).value): c for c in range(1, ws.max_column + 1)}
    for nm in datetime_cols:
        if nm not in hdr:
            continue
        c_idx = hdr[nm]
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=c_idx).number_format = "yyyy-mm-dd hh:mm"
    if "WORK_DATE" in hdr:
        c_idx = hdr["WORK_DATE"]
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=c_idx).number_format = "yyyy-mm-dd"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{max(1, ws.max_row)}"

    # conservative autosize
    for c_idx in range(1, ws.max_column + 1):
        max_len = 0
        for r in range(1, min(ws.max_row, 4000) + 1):
            txt = s(ws.cell(row=r, column=c_idx).value)
            if txt:
                max_len = max(max_len, len(txt.split("\n")[0]))
            if max_len > 70:
                break
        ws.column_dimensions[get_column_letter(c_idx)].width = max(10, min(40, max_len + 2))


def _check_integrity(seg_detail: pd.DataFrame, plan_demand: pd.DataFrame, grid_qty_total: float) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    total_seg_qty = float(pd.to_numeric(seg_detail["SEG_QTY"], errors="coerce").fillna(0.0).sum())
    out["total_seg_qty"] = total_seg_qty
    out["grid_qty_total"] = float(grid_qty_total)
    out["grid_qty_diff"] = float(grid_qty_total - total_seg_qty)

    scheduled = plan_demand.copy()
    if "IS_SCHEDULED" in scheduled.columns:
        scheduled = scheduled[scheduled["IS_SCHEDULED"].map(_truthy)]
    if "ORDER_QTY" in scheduled.columns and "DEMAND_ID" in scheduled.columns:
        scheduled["DEMAND_ID"] = scheduled["DEMAND_ID"].map(_norm_id)
        scheduled["ORDER_QTY"] = pd.to_numeric(scheduled["ORDER_QTY"], errors="coerce").fillna(0.0)
        demand_qty = seg_detail.groupby("DEMAND_ID", dropna=False)["SEG_QTY"].sum()
        mismatch: List[Dict[str, Any]] = []
        for _, r in scheduled[["DEMAND_ID", "ORDER_QTY"]].drop_duplicates().iterrows():
            did = _norm_id(r["DEMAND_ID"])
            expected = float(r["ORDER_QTY"])
            got = float(demand_qty.get(did, 0.0))
            if abs(expected - got) > 1e-6:
                mismatch.append({"demand_id": did, "order_qty": expected, "seg_qty_sum": got, "diff": got - expected})
        out["demand_qty_mismatch_count"] = int(len(mismatch))
        out["demand_qty_mismatches_top20"] = mismatch[:20]
        out["scheduled_order_qty_total"] = float(scheduled["ORDER_QTY"].sum())
    else:
        out["demand_qty_mismatch_count"] = None
        out["scheduled_order_qty_total"] = None
    return out


def _write_ops_summary_sheet(
    wb: Workbook,
    *,
    meta: Dict[str, str],
    seg_detail: pd.DataFrame,
    plan_demand: pd.DataFrame,
    checks: Dict[str, Any],
    cell_qty_unit: QtyUnit,
) -> None:
    ws = ensure_sheet(wb, "00_운영요약", index=0)

    seg = seg_detail.copy()
    seg_qty_series = seg["SEG_QTY"] if "SEG_QTY" in seg.columns else pd.Series(index=seg.index, dtype="float")
    dur_series = seg["DUR_MIN"] if "DUR_MIN" in seg.columns else pd.Series(index=seg.index, dtype="float")
    setup_series = seg["SETUP_IN_MIN"] if "SETUP_IN_MIN" in seg.columns else pd.Series(index=seg.index, dtype="float")
    seg["SEG_QTY"] = pd.to_numeric(seg_qty_series, errors="coerce").fillna(0.0)
    seg["DUR_MIN"] = pd.to_numeric(dur_series, errors="coerce").fillna(0.0)
    seg["SETUP_IN_MIN"] = pd.to_numeric(setup_series, errors="coerce").fillna(0.0)

    unscheduled_count = 0
    total_demand_count = int(len(plan_demand))
    scheduled_demand_count = 0
    if "IS_SCHEDULED" in plan_demand.columns:
        scheduled_series = plan_demand["IS_SCHEDULED"].map(_truthy)
        scheduled_demand_count = int(scheduled_series.sum())
        unscheduled_count = int((~scheduled_series).sum())

    qty_total = float(seg["SEG_QTY"].sum())
    qty_unit_label = "병"
    if cell_qty_unit == "case":
        qty_unit_label = "박스"
    elif cell_qty_unit == "both":
        qty_unit_label = "병(기본)"

    grid_diff = float(checks.get("grid_qty_diff", 0.0) or 0.0)
    mismatch_cnt = int(checks.get("demand_qty_mismatch_count", 0) or 0)
    integrity_ok = abs(grid_diff) <= 1e-6 and mismatch_cnt == 0

    rows = [
        {"항목": "RUN_ID", "값": s(meta.get("RUN_ID")), "설명": "이번 실행 식별자"},
        {"항목": "SCENARIO", "값": s(meta.get("SCENARIO")), "설명": "적용 시나리오"},
        {"항목": "계획 시작일", "값": s(meta.get("START_DATE")), "설명": "계획 기간 시작"},
        {"항목": "계획 종료일", "값": s(meta.get("END_DATE")), "설명": "계획 기간 종료"},
        {"항목": "총 작업 수", "값": int(len(seg)), "설명": "배정된 세그먼트 개수"},
        {"항목": "전체 수요 수", "값": total_demand_count, "설명": "계획 대상 전체 수요 개수"},
        {"항목": "배정 수요 수", "값": scheduled_demand_count, "설명": "작업으로 실제 배정된 수요 개수"},
        {"항목": "미배정 수요 수", "값": unscheduled_count, "설명": "0이 아니면 PLAN/INFEASIBLE 점검 필요"},
        {"항목": f"총 생산 수량({qty_unit_label})", "값": int(round(qty_total)), "설명": "세그먼트 기준 합계"},
        {"항목": "총 RUN 시간(분)", "값": int(round(float(seg["DUR_MIN"].sum()))), "설명": "실생산 시간 합계"},
        {"항목": "총 SETUP 시간(분)", "값": int(round(float(seg["SETUP_IN_MIN"].sum()))), "설명": "전환/세척 포함 준비시간"},
        {"항목": "사용 라인 수", "값": int(seg["LINE_ID"].nunique()) if "LINE_ID" in seg.columns else 0, "설명": "실ZERO 사용된 라인 수"},
        {"항목": "사용 품목 수", "값": int(seg["PRODUCT_ID"].nunique()) if "PRODUCT_ID" in seg.columns else 0, "설명": "실ZERO 생산된 품목 수"},
        {"항목": "무결성 점검", "값": "PASS" if integrity_ok else "FAIL", "설명": f"grid_qty_diff={grid_diff:.3f}, mismatch={mismatch_cnt}"},
        {"항목": "권장 확인 시트", "값": "01_주간생산계획 / 02_라인별작업지시 / 03_품목요약", "설명": "운영자 상세 검토 순서"},
    ]

    write_df(
        ws,
        pd.DataFrame(rows),
        freeze_panes="A2",
        max_col_width=52,
        autofilter=False,
    )
    _apply_table_format(ws)


def generate_ops_plan_xlsx(
    plan_xlsx_path: str,
    out_ops_xlsx_path: str,
    *,
    ssot_xlsx_path: Optional[str] = None,
    gate_json_path: Optional[str] = None,
    run_manifest_json_path: Optional[str] = None,
    line_display_map_path: Optional[str] = None,
    include_inactive_lines_in_columns: bool = False,
    fail_on_ops_report: bool = False,
    max_cell_items: int = 6,
    cell_qty_unit: QtyUnit = "bottle",
    include_hidden_meta: bool = True,
) -> str:
    try:
        if cell_qty_unit not in {"bottle", "case", "both"}:
            raise ValueError(f"OPS_REPORT_INVALID_CELL_QTY_UNIT:{cell_qty_unit}")

        seg, split, plan_demand, plan_staff, meta = _read_required_sheets(plan_xlsx_path)
        seg_detail = _build_seg_detail(seg, split, plan_demand)
        ssot_maps = _load_ssot_maps(ssot_xlsx_path)
        line_display_map = _load_line_display_map(line_display_map_path)

        line_ids, display_name_by_line = _collect_line_columns(
            seg_detail,
            ssot_maps,
            line_display_map,
            include_inactive_lines_in_columns=bool(include_inactive_lines_in_columns),
        )
        dates = _date_window(seg_detail, meta)
        staff_by_date = _build_staff_headcount_by_date(plan_staff, seg_detail)
        grid_df, grid_qty_total = _build_grid_rows(
            seg_detail,
            dates=dates,
            line_ids=line_ids,
            display_name_by_line=display_name_by_line,
            ssot_maps=ssot_maps,
            max_cell_items=int(max(1, max_cell_items)),
            cell_qty_unit=cell_qty_unit,
            staff_by_date=staff_by_date,
        )
        line_headers = [display_name_by_line.get(lid, lid) for lid in line_ids]

        line_df = _build_line_instruction(seg_detail, ssot_maps, display_name_by_line, cell_qty_unit=cell_qty_unit)
        product_df = _build_product_summary(seg_detail, ssot_maps, plan_demand)

        checks = _check_integrity(seg_detail, plan_demand, grid_qty_total)
        check_fail = False
        if abs(float(checks.get("grid_qty_diff", 0.0))) > 1e-6:
            check_fail = True
        if checks.get("demand_qty_mismatch_count"):
            check_fail = True
        if check_fail and fail_on_ops_report:
            raise ValueError(f"OPS_REPORT_INTEGRITY_FAIL:{json.dumps(checks, ensure_ascii=False)}")

        wb = Workbook()
        wb.remove(wb.active)

        _write_grid_sheet(wb, grid_df=grid_df, line_headers=line_headers, date_values=dates, index=1)
        _write_ops_summary_sheet(
            wb,
            meta=meta,
            seg_detail=seg_detail,
            plan_demand=plan_demand,
            checks=checks,
            cell_qty_unit=cell_qty_unit,
        )

        ws2 = ensure_sheet(wb, "02_라인별작업지시", index=2)
        write_df(
            ws2,
            line_df,
            freeze_panes="A2",
            max_col_width=48,
            number_formats={"SEG_QTY_BOTTLE": "#,##0", "SEG_QTY_CASE": "#,##0.0", "DUR_MIN": "#,##0"},
        )
        _apply_table_format(ws2, datetime_cols=["START_DT", "END_DT"])

        ws3 = ensure_sheet(wb, "03_품목요약", index=3)
        write_df(
            ws3,
            product_df,
            freeze_panes="A2",
            max_col_width=48,
            number_formats={"TOTAL_QTY_BOTTLE": "#,##0", "TOTAL_TARDINESS_MIN": "#,##0"},
        )
        _apply_table_format(ws3, datetime_cols=["EARLIEST_START_DT", "LATEST_END_DT"])

        if include_hidden_meta:
            ws_meta = ensure_sheet(wb, "__META_RUN", index=4)
            gate = _load_json(gate_json_path)
            manf = _load_json(run_manifest_json_path)
            rows = [
                {"KEY": "GENERATED_AT_UTC", "VALUE": utcnow_iso()},
                {"KEY": "PLAN_XLSX_PATH", "VALUE": os.path.abspath(str(plan_xlsx_path))},
                {"KEY": "OPS_XLSX_PATH", "VALUE": os.path.abspath(str(out_ops_xlsx_path))},
                {"KEY": "PLAN_XLSX_SHA256", "VALUE": file_sha256(str(plan_xlsx_path)) if os.path.exists(plan_xlsx_path) else ""},
                {"KEY": "SSOT_XLSX_PATH", "VALUE": os.path.abspath(str(ssot_xlsx_path)) if ssot_xlsx_path else ""},
                {"KEY": "SSOT_XLSX_SHA256", "VALUE": file_sha256(str(ssot_xlsx_path)) if ssot_xlsx_path and os.path.exists(ssot_xlsx_path) else ""},
                {"KEY": "SCENARIO", "VALUE": s(meta.get("SCENARIO"))},
                {"KEY": "START_DATE", "VALUE": s(meta.get("START_DATE"))},
                {"KEY": "END_DATE", "VALUE": s(meta.get("END_DATE"))},
                {"KEY": "GRID_TOTAL_QTY", "VALUE": s(checks.get("grid_qty_total"))},
                {"KEY": "SEG_TOTAL_QTY", "VALUE": s(checks.get("total_seg_qty"))},
                {"KEY": "GRID_QTY_DIFF", "VALUE": s(checks.get("grid_qty_diff"))},
                {"KEY": "DEMAND_QTY_MISMATCH_COUNT", "VALUE": s(checks.get("demand_qty_mismatch_count"))},
                {"KEY": "GATE_OVERALL_PASS", "VALUE": s(gate.get("overall_pass", gate.get("pass", "")))},
                {"KEY": "GATE_STATUS", "VALUE": s(gate.get("solver_status", gate.get("status", "")))},
                {"KEY": "RUN_ID", "VALUE": s(manf.get("run_id", meta.get("RUN_ID", "")))},
                {"KEY": "SOURCE_MODE", "VALUE": s(meta.get("SOURCE", manf.get("source_mode", "")))},
            ]
            write_df(ws_meta, pd.DataFrame(rows), freeze_panes="A2", max_col_width=80, autofilter=False)
            ws_meta.sheet_state = "hidden"

        os.makedirs(os.path.dirname(os.path.abspath(out_ops_xlsx_path)) or ".", exist_ok=True)
        wb.save(out_ops_xlsx_path)

        # Re-open and enforce hidden sheet state even if writers reorder.
        if include_hidden_meta:
            w2 = load_workbook(out_ops_xlsx_path)
            if "__META_RUN" in w2.sheetnames:
                w2["__META_RUN"].sheet_state = "hidden"
            w2.save(out_ops_xlsx_path)
        return out_ops_xlsx_path
    except Exception:
        if fail_on_ops_report:
            raise
        return ""
