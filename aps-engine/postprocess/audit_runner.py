"""Postprocess: generate human-friendly RICH audit workbook.

The heavy rendering is delegated to tools/aps_plan_audit_report_ssot_rich.py.
This module only builds adapter DataFrames, calls the renderer, and then adds a
few v20-specific/ops-critical sheets with better readability.
"""

from __future__ import annotations

import os
import traceback
from datetime import datetime
from dataclasses import dataclass
from typing import Any, Dict, Optional

import openpyxl
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd

from .audit_adapter import (
    build_audit_frames,
    build_unscheduled_tables,
)
from .xlsx_pretty import ensure_sheet, write_df, write_index_sheet
from . import xlsx_pretty


@dataclass(frozen=True)
class _SheetSpec:
    name: str
    title: str


_INDEX_SHEETS = (
    _SheetSpec("00_월간보드", "월간보드"),
    _SheetSpec("01_개요", "개요"),
    _SheetSpec("02_일별타임테이블", "일별 타임테이블"),
    _SheetSpec("03_작업목록", "작업목록"),
    _SheetSpec("04_이상치_목록", "이상치"),
    _SheetSpec("05_미배정_수요", "미배정 수요"),
    _SheetSpec("05B_불가능_수요", "불가능 수요"),
    _SheetSpec("06_인력_미충족", "인력 미충족"),
    _SheetSpec("08_가중치_점검", "목적함수/가중치"),
)


def _df(rows) -> pd.DataFrame:
    return pd.DataFrame(rows or [])


def _get_data(data: Any, key: str, default: Any = "") -> Any:
    try:
        return data.get(key, default) if isinstance(data, dict) else getattr(data, key, default)
    except Exception:
        return default


def _norm_path(p: Any) -> str:
    try:
        return os.path.abspath(str(p)) if p else ""
    except Exception:
        return ""


def _build_weights_df(result: Dict[str, Any]) -> pd.DataFrame:
    weights = (result or {}).get("objective_weights", {}) or {}
    rows = []
    for k, v in weights.items():
        try:
            vv = float(v)
        except Exception:
            vv = v
        rows.append({"TERM": k, "WEIGHT": vv})
    df = pd.DataFrame(rows).sort_values("TERM") if rows else pd.DataFrame(columns=["TERM", "WEIGHT"])
    all_zero = False
    try:
        all_zero = (len(df) > 0) and (pd.to_numeric(df["WEIGHT"], errors="coerce").fillna(0) == 0).all()
    except Exception:
        all_zero = False
    if all_zero:
        df = pd.concat(
            [
                pd.DataFrame(
                    [
                        {
                            "TERM": "!!CRITICAL!!",
                            "WEIGHT": "ALL_WEIGHTS_ARE_ZERO -> 'do nothing' becomes optimal",
                        }
                    ]
                ),
                df,
            ],
            ignore_index=True,
        )
    return df


def _build_daily_timetable(plan_gantt: pd.DataFrame, ssot) -> pd.DataFrame:
    if plan_gantt is None or plan_gantt.empty:
        return pd.DataFrame(
            columns=[
                "WORK_DATE",
                "LINE_ID",
                "START_HHMM",
                "END_HHMM",
                "TIME_RANGE",
                "PRODUCT_NAME",
                "PRODUCT_ID",
                "DEMAND_ID",
                "JOB_IDX",
                "RUN_MIN",
                "SETUP_MIN",
                "TOTAL_OCCUPY_MIN",
                "QTY_BOTTLE",
                "REQUIRED_CREW",
            ]
        )

    df = plan_gantt.copy()
    df["START_DT"] = pd.to_datetime(df.get("START_DATE"), errors="coerce")
    df["END_DT"] = pd.to_datetime(df.get("END_DATE"), errors="coerce")
    df["WORK_DATE"] = df["START_DT"].dt.date
    df["START_HHMM"] = df["START_DT"].dt.strftime("%H:%M")
    df["END_HHMM"] = df["END_DT"].dt.strftime("%H:%M")
    df["TIME_RANGE"] = df["START_HHMM"].fillna("") + "~" + df["END_HHMM"].fillna("")

    # Optional SSOT enrichment (product name)
    prod_name = {}
    try:
        if ssot is not None and hasattr(ssot, "product") and ssot.product is not None and not ssot.product.empty:
            if "PRODUCT_ID" in ssot.product.columns:
                name_col = "PRODUCT_NAME_KO" if "PRODUCT_NAME_KO" in ssot.product.columns else None
                if name_col:
                    prod_name = dict(zip(ssot.product["PRODUCT_ID"].astype(str), ssot.product[name_col].astype(str)))
    except Exception:
        prod_name = {}

    df["PRODUCT_ID"] = df.get("PRODUCT_ID").astype(str)
    df["PRODUCT_NAME"] = df["PRODUCT_ID"].map(prod_name).fillna("")

    out = pd.DataFrame(
        {
            "WORK_DATE": df["WORK_DATE"],
            "LINE_ID": df.get("LINE_ID"),
            "START_HHMM": df["START_HHMM"],
            "END_HHMM": df["END_HHMM"],
            "TIME_RANGE": df["TIME_RANGE"],
            "PRODUCT_NAME": df["PRODUCT_NAME"],
            "PRODUCT_ID": df["PRODUCT_ID"],
            "DEMAND_ID": df.get("DEMAND_ID"),
            "JOB_IDX": df.get("JOB_IDX"),
            "RUN_MIN": df.get("RUN_MIN"),
            "SETUP_MIN": df.get("SETUP_MIN"),
            "TOTAL_OCCUPY_MIN": df.get("TOTAL_LINE_OCCUPY_MIN"),
            "QTY_BOTTLE": df.get("QTY_BOTTLE"),
            "REQUIRED_CREW": df.get("REQUIRED_CREW"),
        }
    )
    out = out.sort_values(["WORK_DATE", "LINE_ID", "START_HHMM", "END_HHMM"], kind="mergesort")
    return out


def _build_staff_missing_df(staff_rows, start_date, data: Optional[Dict[str, Any]] = None) -> pd.DataFrame:
    df = _df(staff_rows)
    if df.empty or "ASSIGN_STATUS" not in df.columns:
        return pd.DataFrame(
            columns=[
                "WORK_DATE",
                "LINE_ID",
                "REQUIRED_HEADCOUNT",
                "QUALIFIED_AVAILABLE",
                "SHORTAGE",
                "START_HHMM",
                "END_HHMM",
                "SEGMENT_ID",
                "DEMAND_ID",
                "SEAT_TYPE_CODE",
            ]
        )

    # time columns
    try:
        start_dt = pd.to_datetime(start_date)
        df["START_DT"] = start_dt + pd.to_timedelta(pd.to_numeric(df.get("START_MIN", 0), errors="coerce").fillna(0).astype(int), unit="m")
        df["END_DT"] = start_dt + pd.to_timedelta(pd.to_numeric(df.get("END_MIN", 0), errors="coerce").fillna(0).astype(int), unit="m")
        df["WORK_DATE"] = df["START_DT"].dt.date
        df["START_HHMM"] = df["START_DT"].dt.strftime("%H:%M")
        df["END_HHMM"] = df["END_DT"].dt.strftime("%H:%M")
    except Exception:
        df["WORK_DATE"] = None
        df["START_HHMM"] = None
        df["END_HHMM"] = None

    grp_cols = ["WORK_DATE", "START_HHMM", "END_HHMM", "LINE_ID", "SEGMENT_ID", "DEMAND_ID", "SEAT_TYPE_CODE"]
    for c in grp_cols:
        if c not in df.columns:
            df[c] = None

    df["ASSIGN_STATUS"] = df["ASSIGN_STATUS"].astype(str).str.upper()
    req = df.groupby(grp_cols, dropna=False).size().reset_index(name="REQUIRED_HEADCOUNT")
    miss = df[df["ASSIGN_STATUS"] == "MISSING"].groupby(grp_cols, dropna=False).size().reset_index(name="SHORTAGE")
    out = req.merge(miss, on=grp_cols, how="left")
    out["SHORTAGE"] = pd.to_numeric(out["SHORTAGE"], errors="coerce").fillna(0).astype(int)

    # Qualified available (line-level)
    qual_by_line = {}
    if data is not None:
        q = data.get("qual_by_line_seat") or {}
        for (ln, _seat), quals in q.items():
            if not ln:
                continue
            for rec in quals:
                sid = str(rec.get("STAFF_ID") or "").strip()
                if sid:
                    qual_by_line.setdefault(str(ln), set()).add(sid)

    def _qual_cnt(ln: Any) -> Optional[int]:
        key = str(ln)
        if key in qual_by_line:
            return int(len(qual_by_line[key]))
        return None

    out["QUALIFIED_AVAILABLE"] = out["LINE_ID"].map(_qual_cnt)

    out = out.sort_values(["WORK_DATE", "LINE_ID", "START_HHMM"], kind="mergesort")
    return out


def _compute_global_peak(plan_gantt: pd.DataFrame, *, include_setup: bool = False) -> Tuple[int, Optional[str]]:
    if plan_gantt is None or plan_gantt.empty:
        return 0, None
    df = plan_gantt.copy()
    if "REQUIRED_CREW" not in df.columns:
        return 0, None
    df["CREW"] = pd.to_numeric(df["REQUIRED_CREW"], errors="coerce").fillna(0).astype(int)
    df = df[df["CREW"] > 0]
    if df.empty:
        return 0, None
    df["START_DATE"] = pd.to_datetime(df.get("START_DATE"), errors="coerce")
    df["END_DATE"] = pd.to_datetime(df.get("END_DATE"), errors="coerce")
    if include_setup:
        df["SETUP_MIN"] = pd.to_numeric(df.get("SETUP_MIN"), errors="coerce").fillna(0).astype(int)
        df["OCC_START"] = df["START_DATE"] - pd.to_timedelta(df["SETUP_MIN"], unit="m")
    else:
        df["OCC_START"] = df["START_DATE"]

    events: List[Tuple[pd.Timestamp, int]] = []
    for _, r in df.iterrows():
        st = r.get("OCC_START")
        en = r.get("END_DATE")
        crew = int(r.get("CREW", 0))
        if pd.isna(st) or pd.isna(en) or crew <= 0:
            continue
        events.append((pd.to_datetime(st), crew))
        events.append((pd.to_datetime(en), -crew))

    if not events:
        return 0, None

    cur = 0
    peak = 0
    peak_time: Optional[pd.Timestamp] = None
    for t, delta in sorted(events, key=lambda x: (x[0], 0 if x[1] < 0 else 1)):
        cur += int(delta)
        if cur > peak:
            peak = cur
            peak_time = t
    return int(peak), (peak_time.isoformat(sep=" ") if peak_time is not None else None)


# =============================================================================
# v20 추가 시트(현업 운영용) - rich 보고서 생성 후 후처리
# =============================================================================
def add_v20_custom_sheets(
    *,
    report_out_path: str,
    result: Mapping[str, Any],
    data: Mapping[str, Any],
    plan_gantt: pd.DataFrame,
    show_raw: bool,
    ssot: Any | None = None,
) -> None:
    """Append v20-specific operational sheets to the rich report workbook.

    This function must never raise exceptions that would break the core solve.
    Any error here should be handled by the caller (safe_generate_rich_report).
    """
    wb = load_workbook(report_out_path)

    # 02_일별타임테이블
    try:
        df_tt = _build_daily_timetable(plan_gantt, ssot=ssot)
        ws = ensure_sheet(wb, "02_일별타임테이블")
        write_df(ws, df_tt, freeze_panes="A2")
    except Exception as e:
        ws = ensure_sheet(wb, "02_일별타임테이블")
        ws["A1"] = f"FAILED: {type(e).__name__}: {e}"

    # 05_미배정/불가능 수요
    try:
        unscheduled_df, infeasible_df = build_unscheduled_tables(result)
        ws = ensure_sheet(wb, "05_미배정_수요")
        write_df(ws, unscheduled_df, freeze_panes="A2")
        ws = ensure_sheet(wb, "05B_불가능_수요")
        write_df(ws, infeasible_df, freeze_panes="A2")
    except Exception as e:
        ws = ensure_sheet(wb, "05_미배정_수요")
        ws["A1"] = f"FAILED: {type(e).__name__}: {e}"
        ws = ensure_sheet(wb, "05B_불가능_수요")
        ws["A1"] = f"FAILED: {type(e).__name__}: {e}"

    # 06_인력_미충족 (현재는 집계/슬롯 기반; Seat-level은 Phase2)
    try:
        staff_rows = (result or {}).get("staff_shortage_rows") or (result or {}).get("staff_rows") or []
        start_date = _get_data(data, "start_date") or _get_data(data, "horizon_start") or ""
        df_staff = _build_staff_missing_df(staff_rows, start_date=start_date, data=dict(data))
        ws = ensure_sheet(wb, "06_인력_미충족")
        write_df(ws, df_staff, freeze_panes="A2")
    except Exception as e:
        ws = ensure_sheet(wb, "06_인력_미충족")
        ws["A1"] = f"FAILED: {type(e).__name__}: {e}"

    # 08_가중치_점검
    try:
        df_w = _build_weights_df(result)
        ws = ensure_sheet(wb, "08_가중치_점검")
        write_df(ws, df_w, freeze_panes="A2")
    except Exception as e:
        ws = ensure_sheet(wb, "08_가중치_점검")
        ws["A1"] = f"FAILED: {type(e).__name__}: {e}"

    # RAW 시트 추가(선택) - 기존 RAW_* 외에 v20 출력 원본을 숨김으로 첨부
    if show_raw:
        try:
            xls = pd.ExcelFile(_norm_path(_get_data(data, "raw_out_path") or ""))
        except Exception:
            xls = None

    # 간단한 숨김 처리: rich 쪽 RAW_*는 render_workbook에서 hide_raw로 제어.
    # 여기서는 별도 RAW 추가는 하지 않음(필요하면 확장).

    wb.save(report_out_path)



def generate_rich_report(
    *,
    result: Dict[str, Any],
    data: Dict[str, Any],
    raw_out_path: str,
    report_out_path: str,
    ssot_path: Optional[str],
    scenario_id: str,
    plan_month: Optional[str] = None,
    include_raw: bool = True,
    show_raw: bool = False,
) -> str:
    """Generate enterprise-friendly (rich) audit workbook.

    Design principles:
      - This is an optional action layer. Failures must NOT break the core solver output.
      - Use the same pipeline as tools/aps_plan_audit_report_ssot_rich.py (authoritative).
      - Keep v20-specific extra sheets (unscheduled/infeasible/staff/weights) as additive.
    """

    # Lazy import to avoid import-time coupling between core solver and postprocess.
    from ..tools import aps_plan_audit_report_ssot_rich as rich

    # ---- 1) Build canonical raw frames from v20 solver outputs (adapter layer) ----
    frames = build_audit_frames(result=result, data=data)
    plan_gantt = frames["plan_gantt"]
    daily_segments = frames["daily_segments"]
    daily_line_summary = frames["daily_line_summary"]
    calendar_qc = frames["calendar_qc"]

    # ---- 2) Load SSOT (optional) ----
    ssot = None
    if ssot_path:
        ssot = rich.load_ssot_full(ssot_path, scenario_id=scenario_id)

    # ---- 3) Run the rich-tool pipeline (mirrors rich.main()) ----
    plan_demand = pd.DataFrame(result.get("plan_rows") or [])
    job_table = rich.build_job_table(plan_gantt, ssot, plan_demand=plan_demand if not plan_demand.empty else None)

    # SSOT setup expectation audit (optional)
    if ssot is not None and not plan_gantt.empty:
        exp = rich.compute_expected_setup(plan_gantt, ssot)
        if not exp.empty and "JOB_IDX" in job_table.columns and "JOB_IDX" in exp.columns:
            job_table = job_table.merge(exp, on="JOB_IDX", how="left")
            if "SSOT_SETUP_EXPECT_MIN" in job_table.columns:
                job_table["SSOT_예상셋업(분)"] = job_table["SSOT_SETUP_EXPECT_MIN"]
                # If SETUP(분) exists, compute diff; else leave NaN.
                if "SETUP(분)" in job_table.columns:
                    job_table["셋업차이(APS-SSOT)"] = (
                        pd.to_numeric(job_table["SETUP(분)"], errors="coerce")
                        - pd.to_numeric(job_table["SSOT_SETUP_EXPECT_MIN"], errors="coerce")
                    )

    # Attach line master + calendar view (optional)
    if ssot is not None and not daily_line_summary.empty:
        daily_line_with_cal = rich.attach_calendar_to_daily_line(
            daily_line_summary,
            ssot,
            line_master=ssot.line,
        )
    else:
        daily_line_with_cal = daily_line_summary.copy()

    # Anomaly detection
    job_full, anomalies = rich.detect_anomalies(job_table, daily_line_with_cal, calendar_qc, ssot)

    # Monthly product summary (optional plan_month filter)
    prod_summary = rich.build_monthly_product_summary(job_full, plan_month=plan_month)

    # Monthly board (requires daily aggregates; pass None if missing)
    board_df, board_sev = rich.build_monthly_board(
        job_full,
        daily_segments if not daily_segments.empty else None,
        daily_line_summary if not daily_line_summary.empty else None,
        daily_line_with_cal if not daily_line_with_cal.empty else None,
    )

    # SSOT slices (for reference tabs)
    if ssot is not None and not job_full.empty and "시작일" in job_full.columns and "종료일" in job_full.columns:
        date_min = pd.to_datetime(job_full["시작일"], errors="coerce").min()
        date_max = pd.to_datetime(job_full["종료일"], errors="coerce").max()
        ssot_slices = rich.build_ssot_slices(job_full, ssot, date_min=date_min, date_max=date_max)
    else:
        ssot_slices = None

    overview = {
        "report_generated_at": datetime.now().isoformat(timespec="seconds"),
        "plan_month": plan_month or "",
        "scenario": scenario_id,
        "horizon_start": str(data.get("start_date", "")),
        "horizon_end": str(data.get("end_date", "")),
        "jobs": int(len(job_full)) if isinstance(job_full, pd.DataFrame) else 0,
        "anomalies": int(len(anomalies)) if isinstance(anomalies, pd.DataFrame) else 0,
        "unscheduled_demands": int(len(result.get("unscheduled_demands", []))),
        "infeasible_demands": int(len(result.get("infeasible_demands", []))),
    }
    try:
        peak_prod, _ = _compute_global_peak(plan_gantt, include_setup=False)
        peak_occ, peak_ts = _compute_global_peak(plan_gantt, include_setup=True)
        overview["PEAK_CREW_GLOBAL_PROD"] = int(peak_prod)
        overview["PEAK_CREW_GLOBAL_PROD_SETUP"] = int(peak_occ)
        overview["PEAK_CREW_GLOBAL_PROD_SETUP_AT"] = peak_ts or ""
    except Exception:
        overview["PEAK_CREW_GLOBAL_PROD"] = ""
        overview["PEAK_CREW_GLOBAL_PROD_SETUP"] = ""
        overview["PEAK_CREW_GLOBAL_PROD_SETUP_AT"] = ""

    raw_sheets = None
    if include_raw:
        raw_sheets = {
            "PLAN_GANTT": plan_gantt,
            "PLAN_DEMAND": plan_demand,
            "DAILY_SEGMENTS": daily_segments,
            "DAILY_LINE_SUMMARY": daily_line_summary,
            "CALENDAR_QC": calendar_qc,
        }

    rich.render_workbook(
        output_path=report_out_path,
        board_df=board_df,
        board_sev=board_sev,
        overview=overview,
        prod_summary=prod_summary,
        job_full=job_full,
        anomalies=anomalies,
        daily_line_view=daily_line_with_cal,
        calendar_qc=calendar_qc,
        ssot_slices=ssot_slices,
        raw_sheets=raw_sheets,
        hide_raw=(not show_raw),
    )

    # ---- 4) Add v20 custom sheets (non-breaking, additive) ----
    add_v20_custom_sheets(
        report_out_path=report_out_path,
        result=result,
        data=data,
        plan_gantt=plan_gantt,
        show_raw=show_raw,
        ssot=ssot,
    )

    # ---- 5) Pretty formatting ----
    try:
        xlsx_pretty.apply_pretty_formatting(report_out_path, index_sheets=_INDEX_SHEETS)
    except Exception as e:
        # Formatting must never break the report artifact.
        print(f"RICH_REPORT_PRETTY_FAILED: {e}")

    return report_out_path


def _write_fallback_rich_report(
    report_out_path: str,
    *,
    result: Dict[str, Any],
    data: Dict[str, Any],
    raw_out_path: str,
    ssot_path: Optional[str],
    scenario_id: Optional[str],
    plan_month: Optional[str],
    error_text: str,
) -> None:
    """Create a minimal, always-available rich report workbook.

    목적:
      - tools/aps_plan_audit_report_ssot_rich.py가 예외로 중단돼도,
        운영/분석자가 즉시 확인 가능한 최소 시트(작업목록/미배정/가중치/에러로그)를
        항상 생성한다.
    """
    wb = openpyxl.Workbook()
    try:
        wb.remove(wb.active)
    except Exception:
        pass

    # 01_개요
    summary = {
        "STATUS": "FALLBACK",
        "SCENARIO": scenario_id or "",
        "PLAN_MONTH": plan_month or "",
        "HORIZON_START": str(data.get("start_date", "")),
        "HORIZON_END": str(data.get("end_date", "")),
        "RAW_OUT_PATH": raw_out_path,
        "SSOT_PATH": ssot_path or "",
    }
    ws = ensure_sheet(wb, "01_개요", index=0)
    write_df(ws, pd.DataFrame([summary]))

    # 03_작업목록 (PLAN_GANTT)
    try:
        plan_gantt = build_plan_gantt(result, data)
    except Exception:
        plan_gantt = pd.DataFrame()
    ws = ensure_sheet(wb, "03_작업목록", index=1)
    write_df(ws, plan_gantt)

    # 05_미배정 / 05B_불가능
    try:
        unscheduled, infeasible = build_unscheduled_tables(result)
    except Exception:
        unscheduled, infeasible = (pd.DataFrame(), pd.DataFrame())
    ws = ensure_sheet(wb, "05_미배정_수요", index=2)
    write_df(ws, unscheduled)
    ws = ensure_sheet(wb, "05B_불가능_수요", index=3)
    write_df(ws, infeasible)

    # 08_가중치_점검
    try:
        weights = _build_weights_df(result)
    except Exception:
        weights = pd.DataFrame()
    ws = ensure_sheet(wb, "08_가중치_점검", index=4)
    write_df(ws, weights)

    # 99_에러로그
    ws = ensure_sheet(wb, "99_RICH_REPORT_ERROR", index=99)
    err_df = pd.DataFrame({"TRACE": error_text.splitlines()})
    write_df(ws, err_df)

    try:
        apply_workbook_style(wb)
    except Exception:
        pass

    wb.save(report_out_path)

def safe_generate_rich_report(
    *,
    result: Dict[str, Any],
    data: Dict[str, Any],
    raw_out_path: str,
    report_out_path: str,
    ssot_path: Optional[str],
    scenario_id: Optional[str],
    plan_month: Optional[str],
    include_raw: bool = False,
    show_raw: bool = False,
) -> bool:
    """Generate rich report; if it fails, write a fallback workbook and return False."""
    try:
        generate_rich_report(
            result=result,
            data=data,
            raw_out_path=raw_out_path,
            report_out_path=report_out_path,
            ssot_path=ssot_path,
            scenario_id=scenario_id,
            plan_month=plan_month,
            include_raw=include_raw,
            show_raw=show_raw,
        )
        print("RICH_REPORT_WRITTEN:", report_out_path)
        return True
    except Exception:
        tb = traceback.format_exc()
        print("RICH_REPORT_FAILED:\n" + tb)
        try:
            _write_fallback_rich_report(
                report_out_path,
                result=result,
                data=data,
                raw_out_path=raw_out_path,
                ssot_path=ssot_path,
                scenario_id=scenario_id,
                plan_month=plan_month,
                error_text=tb,
            )
            print("RICH_REPORT_FALLBACK_WRITTEN:", report_out_path)
        except Exception:
            print("RICH_REPORT_FALLBACK_FAILED:\n" + traceback.format_exc())
        return False
