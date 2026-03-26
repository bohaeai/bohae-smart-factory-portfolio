from __future__ import annotations

import datetime as _dt
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


# -----------------------------------------------------------------------------
# Sheet helpers
# -----------------------------------------------------------------------------

def ensure_sheet(wb: Workbook, name: str, *, index: Optional[int] = None) -> Worksheet:
    """Create a new empty sheet (remove existing one if present)."""
    if name in wb.sheetnames:
        wb.remove(wb[name])
    return wb.create_sheet(title=name, index=index)


# -----------------------------------------------------------------------------
# DataFrame writer (openpyxl-only)
# -----------------------------------------------------------------------------

_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill("solid", fgColor="E7E6E6")  # light gray
_HEADER_ALIGN = Alignment(vertical="center", wrap_text=True)
_CELL_ALIGN = Alignment(vertical="top", wrap_text=True)


def write_df(
    ws: Worksheet,
    df: pd.DataFrame,
    *,
    start_row: int = 1,
    start_col: int = 1,
    freeze_panes: Optional[str] = "A2",
    autofilter: bool = True,
    max_col_width: int = 60,
    number_formats: Optional[Dict[str, str]] = None,
) -> None:
    """Write df into worksheet with a minimal, consistent style.

    - Always writes a header row.
    - If df is empty, writes headers (if any) and a "NO_DATA" marker.
    """
    if number_formats is None:
        number_formats = {}

    if df is None:
        df = pd.DataFrame()

    # Normalize to avoid NaN-looking strings.
    df2 = df.copy()

    # Header
    headers: List[str] = list(df2.columns)
    for j, h in enumerate(headers, start=start_col):
        cell = ws.cell(row=start_row, column=j, value=str(h))
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN

    if df2.empty:
        # Leave a clear marker for human readers.
        ws.cell(row=start_row + 1, column=start_col, value="NO_DATA")
        _post_format(ws, start_row, start_col, len(headers), 1, freeze_panes, autofilter, max_col_width)
        return

    # Body
    for i, row in enumerate(df2.itertuples(index=False), start=start_row + 1):
        for j, val in enumerate(row, start=start_col):
            cell = ws.cell(row=i, column=j, value=_normalize_excel_value(val))
            cell.alignment = _CELL_ALIGN

    # Apply number formats by column header
    if number_formats and headers:
        hdr_to_idx = {h: k for k, h in enumerate(headers, start=start_col)}
        for col_name, fmt in number_formats.items():
            if col_name not in hdr_to_idx:
                continue
            col_idx = hdr_to_idx[col_name]
            for r in range(start_row + 1, start_row + 1 + len(df2)):
                ws.cell(row=r, column=col_idx).number_format = fmt

    _post_format(ws, start_row, start_col, len(headers), len(df2), freeze_panes, autofilter, max_col_width)


def _post_format(
    ws: Worksheet,
    start_row: int,
    start_col: int,
    ncols: int,
    nrows: int,
    freeze_panes: Optional[str],
    autofilter: bool,
    max_col_width: int,
) -> None:
    # Freeze panes
    if freeze_panes:
        ws.freeze_panes = freeze_panes

    # Autofilter (header row)
    if autofilter and ncols > 0:
        # Determine range
        end_row = start_row + nrows
        end_col = start_col + ncols - 1
        ws.auto_filter.ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"

    # Zoom a bit for readability
    ws.sheet_view.zoomScale = 110

    # Autosize columns
    if ncols > 0:
        _autosize_columns(ws, start_row, start_col, ncols, nrows, max_col_width)


def _autosize_columns(
    ws: Worksheet,
    start_row: int,
    start_col: int,
    ncols: int,
    nrows: int,
    max_width: int,
) -> None:
    # Header + body scanning (safe and deterministic)
    for c in range(start_col, start_col + ncols):
        max_len = 0
        for r in range(start_row, start_row + nrows + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        # width heuristic
        width = min(max_len + 2, max_width)
        ws.column_dimensions[get_column_letter(c)].width = max(width, 10)


def _normalize_excel_value(v: Any) -> Any:
    if v is None:
        return None
    # Pandas often gives NaN/NaT
    try:
        if pd.isna(v):
            return None
    except Exception:
        pass
    # Keep datetime/date types (so number_format works)
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v
    return v


# -----------------------------------------------------------------------------
# Index sheet (internal hyperlinks)
# -----------------------------------------------------------------------------


def write_index_sheet(
    wb: Workbook,
    *,
    sheet_name: str,
    links: Sequence[Tuple[str, str]],
    intro_lines: Optional[Sequence[str]] = None,
) -> None:
    """Create a small navigation sheet with internal hyperlinks.

    links: (label, target_sheet_name)
    """
    ws = ensure_sheet(wb, sheet_name, index=0)
    ws.sheet_view.zoomScale = 120

    row = 1
    ws.cell(row=row, column=1, value="RICH REPORT INDEX").font = Font(bold=True, size=14)
    row += 2

    if intro_lines:
        for line in intro_lines:
            ws.cell(row=row, column=1, value=str(line)).alignment = Alignment(wrap_text=True)
            row += 1
        row += 1

    ws.cell(row=row, column=1, value="바로가기").font = Font(bold=True)
    row += 1

    for label, target in links:
        c = ws.cell(row=row, column=1, value=str(label))
        c.font = Font(color="0563C1", underline="single")
        c.hyperlink = f"#'{target}'!A1"
        row += 1

    ws.column_dimensions["A"].width = 50


# =============================================================================
# Optional pretty formatting (non-breaking)
# =============================================================================
from openpyxl.utils import get_column_letter

def _autofit(ws, min_w: int = 9, max_w: int = 70) -> None:
    """Best-effort column autofit based on cell string lengths."""
    try:
        max_col = ws.max_column
        max_row = ws.max_row
    except Exception:
        return
    if not max_col or not max_row:
        return

    for col in range(1, max_col + 1):
        mx = 0
        letter = get_column_letter(col)
        for row in range(1, max_row + 1):
            v = ws.cell(row, col).value
            if v is None:
                continue
            s = str(v)
            # multi-line
            if "\n" in s:
                l = max(len(x) for x in s.splitlines())
            else:
                l = len(s)
            if l > mx:
                mx = l
            # micro-optimization: early stop if already huge
            if mx >= max_w:
                break
        ws.column_dimensions[letter].width = max(min_w, min(max_w, mx + 2))

def apply_pretty_formatting(path: str, index_sheets: Optional[List[str]] = None) -> None:
    """Apply lightweight formatting to a workbook.

    This function is intentionally conservative. It must never raise in normal usage.
    """
    try:
        wb = openpyxl.load_workbook(path)
    except Exception:
        return

    targets = set(index_sheets or [])
    for ws in wb.worksheets:
        if targets and ws.title not in targets:
            continue
        _autofit(ws)

    try:
        wb.save(path)
    except Exception:
        return
